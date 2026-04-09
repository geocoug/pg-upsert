"""Export QA failures as a "fix sheet".

A fix sheet is a user-facing report of violating staging rows, organised
so that each unique staging row appears **once** with a consolidated
``_issues`` column describing every problem found on that row.

Output goes to a user-specified directory.  Three formats are supported:

- **CSV** — one file per table named ``pg_upsert_failures_<table>.csv``
  plus an optional ``pg_upsert_failures_schema.csv`` for schema-level
  issues.
- **JSON** — a single file ``pg_upsert_failures.json`` with the structure
  ``{"<table>": [...], "_schema": [...]}``.
- **XLSX** — a single workbook ``pg_upsert_failures.xlsx`` with one sheet
  per table plus an optional ``_schema`` sheet.  Requires ``openpyxl``.
"""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path
from typing import Any

from .models import QAError, RowViolation, SchemaIssue

logger = logging.getLogger(__name__)

_VALID_FORMATS = ("csv", "json", "xlsx")
_FILE_PREFIX = "pg_upsert_failures"
_ISSUES_COL = "_issues"
_ISSUE_TYPES_COL = "_issue_types"


def export_failures(
    errors: list[QAError],
    directory: str | Path,
    fmt: str = "csv",
) -> Path | None:
    """Write a fix sheet for *errors* into *directory*.

    Args:
        errors: QAError objects from a completed run.  Violations are
            read from ``error.violations`` and schema issues from
            ``error.schema_issues``.
        directory: Output directory (created if it does not exist).
        fmt: One of ``"csv"``, ``"json"``, or ``"xlsx"``.

    Returns:
        The directory path written to, or ``None`` if there were no
        exportable violations.
    """
    fmt = fmt.lower()
    if fmt not in _VALID_FORMATS:
        raise ValueError(
            f"Unsupported export format {fmt!r}. Supported formats: {', '.join(_VALID_FORMATS)}",
        )

    directory = Path(directory)

    # Group violations by table and collect schema issues across all errors.
    fix_sheets: dict[str, list[dict[str, Any]]] = {}
    fix_sheet_columns: dict[str, list[str]] = {}
    schema_rows: list[dict[str, Any]] = []

    # Walk errors, grouping per-table violation lists (not yet deduped).
    per_table_violations: dict[str, list[RowViolation]] = {}
    for err in errors:
        if err.violations:
            per_table_violations.setdefault(err.table, []).extend(err.violations)
        if err.schema_issues:
            for issue in err.schema_issues:
                schema_rows.append(_schema_issue_to_dict(err.table, issue))

    for table, violations in per_table_violations.items():
        rows, columns = _build_fix_sheet(violations)
        if rows:
            fix_sheets[table] = rows
            fix_sheet_columns[table] = columns

    if not fix_sheets and not schema_rows:
        return None

    directory.mkdir(parents=True, exist_ok=True)

    if fmt == "csv":
        _write_csv(fix_sheets, fix_sheet_columns, schema_rows, directory)
    elif fmt == "json":
        _write_json(fix_sheets, schema_rows, directory)
    elif fmt == "xlsx":
        _write_xlsx(fix_sheets, fix_sheet_columns, schema_rows, directory)

    total_rows = sum(len(r) for r in fix_sheets.values())
    logger.info(
        f"Exported {total_rows} failing rows across {len(fix_sheets)} table(s) to {directory} ({fmt})",
    )
    if schema_rows:
        logger.info(f"Exported {len(schema_rows)} schema issue(s)")

    return directory


# ---------------------------------------------------------------------------
# Fix sheet builder
# ---------------------------------------------------------------------------


def _build_fix_sheet(
    violations: list[RowViolation],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Dedupe violations by PK and merge all issues per unique staging row.

    Returns a tuple of ``(rows, columns)``:

    - ``rows``: list of dicts, one per unique violating staging row,
      with table columns plus ``_issues`` and ``_issue_types``.
    - ``columns``: ordered list of column names — PK columns first (as
      discovered from the first violation), then remaining data columns
      from the row in stable order, then ``_issues`` and ``_issue_types``.
    """
    if not violations:
        return [], []

    # Group by pk_values (the dedup key).
    by_pk: dict[tuple, list[RowViolation]] = {}
    for v in violations:
        by_pk.setdefault(v.pk_values, []).append(v)

    # Determine column order from the first violation's row_data.
    # psycopg2 returns columns in SELECT order (table's declared order
    # for SELECT *), so row_data insertion order is the column order.
    first_violation = next(iter(by_pk.values()))[0]
    first_row = first_violation.row_data
    data_columns = list(first_row.keys())
    pk_columns = list(first_violation.pk_columns)
    columns = data_columns + [_ISSUES_COL, _ISSUE_TYPES_COL]

    rows: list[dict[str, Any]] = []
    for _pk, vs in by_pk.items():
        row_data = dict(vs[0].row_data)
        descriptions = sorted({v.description for v in vs if v.description})
        issue_types = sorted({v.issue_type for v in vs if v.issue_type})
        row_data[_ISSUES_COL] = "; ".join(descriptions)
        row_data[_ISSUE_TYPES_COL] = ",".join(issue_types)
        rows.append(row_data)

    # Sort rows for deterministic, human-friendly output. Prefer sorting
    # by PK columns (declared order); fall back to all data columns for
    # tables with no PK. Stringify values so mixed None/int/str types
    # sort without TypeError.
    sort_cols = pk_columns or data_columns
    rows.sort(
        key=lambda r: tuple("" if r.get(c) is None else str(r.get(c)) for c in sort_cols),
    )

    return rows, columns


def _schema_issue_to_dict(table: str, issue: SchemaIssue) -> dict[str, Any]:
    return {
        "table": table,
        "check_type": issue.check_type,
        "column_name": issue.column_name,
        "staging_type": issue.staging_type if issue.staging_type is not None else "",
        "base_type": issue.base_type if issue.base_type is not None else "",
        "description": issue.description,
    }


# ---------------------------------------------------------------------------
# CSV writer
# ---------------------------------------------------------------------------


def _write_csv(
    fix_sheets: dict[str, list[dict[str, Any]]],
    fix_sheet_columns: dict[str, list[str]],
    schema_rows: list[dict[str, Any]],
    directory: Path,
) -> None:
    for table, rows in fix_sheets.items():
        columns = fix_sheet_columns[table]
        path = directory / f"{_FILE_PREFIX}_{table}.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=columns,
                extrasaction="ignore",
                restval="",
            )
            writer.writeheader()
            writer.writerows(rows)

    if schema_rows:
        columns = ["table", "check_type", "column_name", "staging_type", "base_type", "description"]
        path = directory / f"{_FILE_PREFIX}_schema.csv"
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore", restval="")
            writer.writeheader()
            writer.writerows(schema_rows)


# ---------------------------------------------------------------------------
# JSON writer
# ---------------------------------------------------------------------------


def _write_json(
    fix_sheets: dict[str, list[dict[str, Any]]],
    schema_rows: list[dict[str, Any]],
    directory: Path,
) -> None:
    data: dict[str, list[dict[str, Any]]] = dict(fix_sheets)
    if schema_rows:
        data["_schema"] = schema_rows
    path = directory / f"{_FILE_PREFIX}.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, default=str)


# ---------------------------------------------------------------------------
# XLSX writer
# ---------------------------------------------------------------------------


def _write_xlsx(
    fix_sheets: dict[str, list[dict[str, Any]]],
    fix_sheet_columns: dict[str, list[str]],
    schema_rows: list[dict[str, Any]],
    directory: Path,
) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        raise ImportError(
            "XLSX export requires openpyxl. Install with: pip install pg-upsert[xlsx]",
        ) from None

    wb = Workbook()
    # Remove the default blank sheet; we'll create named ones.
    if wb.active is not None:
        wb.remove(wb.active)

    used_sheet_names: set[str] = set()
    for table, rows in fix_sheets.items():
        columns = fix_sheet_columns[table]
        sheet_title = _safe_sheet_name(table, used_sheet_names)
        used_sheet_names.add(sheet_title)
        ws = wb.create_sheet(title=sheet_title)
        ws.append(columns)
        for row in rows:
            ws.append([_xlsx_cell(row.get(c, "")) for c in columns])

    if schema_rows:
        columns = ["table", "check_type", "column_name", "staging_type", "base_type", "description"]
        schema_title = _safe_sheet_name("_schema", used_sheet_names)
        ws = wb.create_sheet(title=schema_title)
        ws.append(columns)
        for row in schema_rows:
            ws.append([_xlsx_cell(row.get(c, "")) for c in columns])

    path = directory / f"{_FILE_PREFIX}.xlsx"
    wb.save(path)


def _safe_sheet_name(table: str, used: set[str]) -> str:
    """Return an Excel-safe (<=31 char) sheet title unique within *used*.

    Excel caps sheet names at 31 characters.  If *table* is longer it is
    truncated; if the truncated name is already in *used*, an incrementing
    suffix (``_2``, ``_3``, ...) is appended so no sheet silently
    overwrites another.  A warning is logged whenever truncation or
    suffixing happens.
    """
    max_len = 31
    base = table[:max_len]
    if base != table:
        logger.warning(
            f"Sheet name truncated from {table!r} to {base!r} (Excel 31-char limit)",
        )
    if base not in used:
        return base
    # Collision — append a numeric suffix, shrinking base as needed.
    for i in range(2, 1000):
        suffix = f"_{i}"
        candidate = f"{base[: max_len - len(suffix)]}{suffix}"
        if candidate not in used:
            logger.warning(
                f"Sheet name collision for {table!r}; using {candidate!r} instead",
            )
            return candidate
    # Extremely unlikely — fall back to base and let openpyxl deal with it.
    return base


def _xlsx_cell(value: Any) -> Any:
    """Coerce values to XLSX-safe types.

    openpyxl accepts str, int, float, bool, datetime, None — anything
    else (Decimal, date, etc.) works via str conversion.
    """
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
