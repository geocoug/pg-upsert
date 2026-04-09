"""Tests for pg_upsert.export — fix sheet model, no database required."""

from __future__ import annotations

import csv
import json
from pathlib import Path

import pytest

from pg_upsert.export import export_failures
from pg_upsert.models import (
    QACheckType,
    QAError,
    RowViolation,
    SchemaIssue,
    TableResult,
    UpsertResult,
)

# ---------------------------------------------------------------------------
# Helpers — build QAError objects with RowViolation payloads
# ---------------------------------------------------------------------------


def _pk_dup_error(table: str = "books") -> QAError:
    """Two staging rows that share PK book_id=101."""
    rows = [
        {"book_id": 101, "title": "Dune", "genre": "sci-fi", "price": 9.99},
        {"book_id": 101, "title": "Dune (reprint)", "genre": "sci-fi", "price": 12.99},
    ]
    return QAError(
        table=table,
        check_type=QACheckType.PRIMARY_KEY,
        details="1 duplicate keys (2 rows)",
        violations=[
            RowViolation(
                pk_values=(101,),
                pk_columns=["book_id"],
                row_data=rows[0],
                issue_type="pk",
                constraint_name="books_pkey",
                description="duplicate PK (book_id)",
            ),
            RowViolation(
                pk_values=(101,),
                pk_columns=["book_id"],
                row_data=rows[1],
                issue_type="pk",
                constraint_name="books_pkey",
                description="duplicate PK (book_id)",
            ),
        ],
    )


def _null_error_book_101(table: str = "books") -> QAError:
    """book_id=101 row with NULL genre (collides with PK dup error above)."""
    return QAError(
        table=table,
        check_type=QACheckType.NULL,
        details="genre (1)",
        violations=[
            RowViolation(
                pk_values=(101,),
                pk_columns=["book_id"],
                row_data={"book_id": 101, "title": "Dune", "genre": None, "price": 9.99},
                issue_type="null",
                issue_column="genre",
                description="NULL in 'genre'",
            ),
        ],
    )


def _null_error_book_205(table: str = "books") -> QAError:
    return QAError(
        table=table,
        check_type=QACheckType.NULL,
        details="title (1)",
        violations=[
            RowViolation(
                pk_values=(205,),
                pk_columns=["book_id"],
                row_data={"book_id": 205, "title": None, "genre": "sci-fi", "price": 14.99},
                issue_type="null",
                issue_column="title",
                description="NULL in 'title'",
            ),
        ],
    )


def _fk_error_authors(table: str = "authors") -> QAError:
    return QAError(
        table=table,
        check_type=QACheckType.FOREIGN_KEY,
        details="fk_publisher (1)",
        violations=[
            RowViolation(
                pk_values=(42,),
                pk_columns=["author_id"],
                row_data={"author_id": 42, "name": "Bob", "publisher_id": 999},
                issue_type="fk",
                issue_column="publisher_id",
                constraint_name="fk_publisher",
                description="FK violation: publisher_id -> public.publishers(publisher_id)",
            ),
        ],
    )


def _ck_error_price(table: str = "books") -> QAError:
    return QAError(
        table=table,
        check_type=QACheckType.CHECK_CONSTRAINT,
        details="price_positive (1)",
        violations=[
            RowViolation(
                pk_values=(300,),
                pk_columns=["book_id"],
                row_data={"book_id": 300, "title": "Free", "genre": "fiction", "price": -1},
                issue_type="ck",
                constraint_name="price_positive",
                description="check 'price_positive' failed",
            ),
        ],
    )


def _unique_error(table: str = "books") -> QAError:
    return QAError(
        table=table,
        check_type=QACheckType.UNIQUE,
        details="uq_isbn (1 duplicates, 2 rows)",
        violations=[
            RowViolation(
                pk_values=(400,),
                pk_columns=["book_id"],
                row_data={"book_id": 400, "title": "A", "genre": "x", "price": 1.0},
                issue_type="unique",
                issue_column="isbn",
                constraint_name="uq_isbn",
                description="duplicate unique (isbn)",
            ),
        ],
    )


def _column_missing_error(table: str = "genres") -> QAError:
    return QAError(
        table=table,
        check_type=QACheckType.COLUMN_EXISTENCE,
        details="description",
        schema_issues=[
            SchemaIssue(
                check_type="column",
                column_name="description",
                description="missing column 'description'",
            ),
        ],
    )


def _type_mismatch_error(table: str = "genres") -> QAError:
    return QAError(
        table=table,
        check_type=QACheckType.TYPE_MISMATCH,
        details="priority (int4 → text)",
        schema_issues=[
            SchemaIssue(
                check_type="type",
                column_name="priority",
                staging_type="int4",
                base_type="text",
                description="type mismatch: 'priority' is int4 in staging, text in base",
            ),
        ],
    )


def _read_csv(path: Path) -> list[dict]:
    with open(path) as f:
        return list(csv.DictReader(f))


# ---------------------------------------------------------------------------
# Empty / no-op cases
# ---------------------------------------------------------------------------


class TestEmpty:
    def test_empty_errors_returns_none(self, tmp_path: Path):
        assert export_failures([], tmp_path / "out") is None

    def test_errors_with_no_violations_returns_none(self, tmp_path: Path):
        err = QAError(table="t", check_type=QACheckType.NULL, details="col (1)")
        assert export_failures([err], tmp_path / "out") is None

    def test_unsupported_format(self, tmp_path: Path):
        with pytest.raises(ValueError, match="Unsupported export format"):
            export_failures([_pk_dup_error()], tmp_path / "out", fmt="parquet")


# ---------------------------------------------------------------------------
# CSV format — one file per table
# ---------------------------------------------------------------------------


class TestExportCsv:
    def test_single_table_file(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error()], out, fmt="csv")
        file = out / "pg_upsert_failures_books.csv"
        assert file.exists()
        rows = _read_csv(file)
        # PK dedup: two violations with pk (101,) merge to one row.
        assert len(rows) == 1
        assert rows[0]["book_id"] == "101"
        assert rows[0]["_issues"] == "duplicate PK (book_id)"
        assert rows[0]["_issue_types"] == "pk"

    def test_multi_table(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error(), _fk_error_authors()], out, fmt="csv")
        assert (out / "pg_upsert_failures_books.csv").exists()
        assert (out / "pg_upsert_failures_authors.csv").exists()
        # Each file has only its own columns.
        books = _read_csv(out / "pg_upsert_failures_books.csv")
        authors = _read_csv(out / "pg_upsert_failures_authors.csv")
        assert "publisher_id" not in books[0]
        assert "book_id" not in authors[0]

    def test_pk_based_dedup_across_checks(self, tmp_path: Path):
        """Same staging row failing PK and NULL checks becomes one fix-sheet row."""
        out = tmp_path / "failures"
        export_failures(
            [_pk_dup_error(), _null_error_book_101()],
            out,
            fmt="csv",
        )
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        # book_id=101 appears once (merged) + book_id=101 (reprint) appears once
        # because the dedup key is pk_values=(101,) for both PK duplicate rows.
        # Wait — both PK duplicate rows have the same pk_values (that's why
        # they're duplicates!), so they collapse to one fix-sheet entry.
        assert len(rows) == 1
        # The merged row has both descriptions.
        assert "NULL in 'genre'" in rows[0]["_issues"]
        assert "duplicate PK (book_id)" in rows[0]["_issues"]
        # _issue_types is sorted, comma-joined, deduped.
        assert rows[0]["_issue_types"] == "null,pk"

    def test_distinct_rows_stay_distinct(self, tmp_path: Path):
        """Rows with different PKs stay separate."""
        out = tmp_path / "failures"
        export_failures(
            [_null_error_book_101(), _null_error_book_205()],
            out,
            fmt="csv",
        )
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        assert len(rows) == 2
        pks = {r["book_id"] for r in rows}
        assert pks == {"101", "205"}

    def test_issues_are_sorted_and_deduped(self, tmp_path: Path):
        """Repeated descriptions are deduped and sorted in _issues."""
        out = tmp_path / "failures"
        # Two identical PK duplicate violations — both produce the same
        # description "duplicate PK (book_id)". Should appear once.
        export_failures([_pk_dup_error()], out, fmt="csv")
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        assert rows[0]["_issues"] == "duplicate PK (book_id)"

    def test_schema_file_written(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures(
            [_pk_dup_error(), _column_missing_error(), _type_mismatch_error()],
            out,
            fmt="csv",
        )
        schema_file = out / "pg_upsert_failures_schema.csv"
        assert schema_file.exists()
        rows = _read_csv(schema_file)
        assert len(rows) == 2
        # Schema rows carry the table column since they span tables.
        assert all("table" in r for r in rows)
        assert rows[0]["column_name"] == "description"
        assert rows[1]["staging_type"] == "int4"
        assert rows[1]["base_type"] == "text"

    def test_no_schema_file_when_none(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error()], out, fmt="csv")
        assert not (out / "pg_upsert_failures_schema.csv").exists()

    def test_directory_created_if_missing(self, tmp_path: Path):
        out = tmp_path / "nested" / "deep" / "failures"
        assert not out.exists()
        export_failures([_pk_dup_error()], out, fmt="csv")
        assert out.exists()
        assert (out / "pg_upsert_failures_books.csv").exists()

    def test_mixed_check_types_on_one_table(self, tmp_path: Path):
        """PK dup (101), NULL (101), NULL (205), CK (300) — 3 unique rows."""
        out = tmp_path / "failures"
        export_failures(
            [
                _pk_dup_error(),
                _null_error_book_101(),
                _null_error_book_205(),
                _ck_error_price(),
            ],
            out,
            fmt="csv",
        )
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        assert len(rows) == 3  # 101 (merged), 205, 300
        by_pk = {r["book_id"]: r for r in rows}
        assert set(by_pk.keys()) == {"101", "205", "300"}
        assert "null" in by_pk["101"]["_issue_types"]
        assert "pk" in by_pk["101"]["_issue_types"]
        assert by_pk["205"]["_issue_types"] == "null"
        assert by_pk["300"]["_issue_types"] == "ck"

    def test_rows_sorted_by_pk(self, tmp_path: Path):
        """Fix sheet rows are ordered by PK values, not by arbitrary columns."""
        # Feed violations in out-of-order PK sequence and verify the output
        # comes out sorted by book_id ascending.
        out = tmp_path / "failures"
        export_failures(
            [_ck_error_price(), _null_error_book_101(), _null_error_book_205()],
            out,
            fmt="csv",
        )
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        assert [r["book_id"] for r in rows] == ["101", "205", "300"]

    def test_pk_description_lists_columns(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error()], out, fmt="csv")
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        assert rows[0]["_issues"] == "duplicate PK (book_id)"

    def test_fk_description_references_target_table(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_fk_error_authors()], out, fmt="csv")
        rows = _read_csv(out / "pg_upsert_failures_authors.csv")
        assert rows[0]["_issues"] == ("FK violation: publisher_id -> public.publishers(publisher_id)")

    def test_unique_description_lists_columns(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_unique_error()], out, fmt="csv")
        rows = _read_csv(out / "pg_upsert_failures_books.csv")
        assert rows[0]["_issues"] == "duplicate unique (isbn)"


# ---------------------------------------------------------------------------
# JSON format — single nested file
# ---------------------------------------------------------------------------


class TestExportJson:
    def test_single_table(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error()], out, fmt="json")
        path = out / "pg_upsert_failures.json"
        assert path.exists()
        with open(path) as f:
            data = json.load(f)
        assert "books" in data
        assert len(data["books"]) == 1
        assert data["books"][0]["_issue_types"] == "pk"

    def test_multi_table_nested(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error(), _fk_error_authors()], out, fmt="json")
        with open(out / "pg_upsert_failures.json") as f:
            data = json.load(f)
        assert set(data.keys()) == {"books", "authors"}

    def test_schema_key(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error(), _column_missing_error()], out, fmt="json")
        with open(out / "pg_upsert_failures.json") as f:
            data = json.load(f)
        assert "_schema" in data
        assert len(data["_schema"]) == 1
        assert data["_schema"][0]["column_name"] == "description"
        assert data["_schema"][0]["table"] == "genres"

    def test_no_schema_key_when_none(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_pk_dup_error()], out, fmt="json")
        with open(out / "pg_upsert_failures.json") as f:
            data = json.load(f)
        assert "_schema" not in data

    def test_null_values_preserved(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures([_null_error_book_205()], out, fmt="json")
        with open(out / "pg_upsert_failures.json") as f:
            data = json.load(f)
        # title was NULL in the staging row.
        assert data["books"][0]["title"] is None


# ---------------------------------------------------------------------------
# XLSX format — single workbook with sheets per table
# ---------------------------------------------------------------------------


class TestExportXlsx:
    def test_sheets_per_table(self, tmp_path: Path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        out = tmp_path / "failures"
        export_failures([_pk_dup_error(), _fk_error_authors()], out, fmt="xlsx")
        path = out / "pg_upsert_failures.xlsx"
        assert path.exists()
        wb = load_workbook(path)
        assert "books" in wb.sheetnames
        assert "authors" in wb.sheetnames
        # One header row + one fix-sheet row for books.
        ws = wb["books"]
        assert ws.max_row == 2
        assert ws.cell(1, 1).value == "book_id"

    def test_schema_sheet(self, tmp_path: Path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        out = tmp_path / "failures"
        export_failures([_pk_dup_error(), _type_mismatch_error()], out, fmt="xlsx")
        wb = load_workbook(out / "pg_upsert_failures.xlsx")
        assert "_schema" in wb.sheetnames

    def test_no_schema_sheet_when_none(self, tmp_path: Path):
        pytest.importorskip("openpyxl")
        from openpyxl import load_workbook

        out = tmp_path / "failures"
        export_failures([_pk_dup_error()], out, fmt="xlsx")
        wb = load_workbook(out / "pg_upsert_failures.xlsx")
        assert "_schema" not in wb.sheetnames

    def test_xlsx_missing_openpyxl_error(self, tmp_path: Path, monkeypatch: pytest.MonkeyPatch):
        import pg_upsert.export as export_mod

        real_import = __builtins__.__import__ if hasattr(__builtins__, "__import__") else __import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("No module named 'openpyxl'")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr("builtins.__import__", fake_import)
        with pytest.raises(ImportError, match="pip install pg-upsert"):
            export_mod._write_xlsx(
                {"books": [{"book_id": 1, "_issues": "x", "_issue_types": "pk"}]},
                {"books": ["book_id", "_issues", "_issue_types"]},
                [],
                tmp_path / "out",
            )


# ---------------------------------------------------------------------------
# All check types covered
# ---------------------------------------------------------------------------


class TestAllCheckTypes:
    def test_all_data_check_types_csv(self, tmp_path: Path):
        out = tmp_path / "failures"
        export_failures(
            [
                _pk_dup_error(),
                _null_error_book_205(),
                _fk_error_authors(),
                _unique_error(),
                _ck_error_price(),
            ],
            out,
            fmt="csv",
        )
        # books file has PK, NULL, UNIQUE, CK — 4 distinct PKs (101, 205, 400, 300)
        books = _read_csv(out / "pg_upsert_failures_books.csv")
        assert len(books) == 4
        types = {r["_issue_types"] for r in books}
        assert types == {"pk", "null", "unique", "ck"}
        # authors file has the FK violation
        authors = _read_csv(out / "pg_upsert_failures_authors.csv")
        assert len(authors) == 1
        assert authors[0]["_issue_types"] == "fk"


# ---------------------------------------------------------------------------
# UpsertResult integration
# ---------------------------------------------------------------------------


class TestUpsertResultExport:
    def test_export_via_result(self, tmp_path: Path):
        err = _pk_dup_error()
        tr = TableResult(table_name="books", qa_errors=[err])
        result = UpsertResult(tables=[tr])
        out = tmp_path / "failures"
        result.export_failures(out)
        assert (out / "pg_upsert_failures_books.csv").exists()

    def test_export_via_result_json(self, tmp_path: Path):
        err = _pk_dup_error()
        tr = TableResult(table_name="books", qa_errors=[err])
        result = UpsertResult(tables=[tr])
        out = tmp_path / "failures"
        result.export_failures(out, fmt="json")
        assert (out / "pg_upsert_failures.json").exists()

    def test_to_dict_excludes_violations(self):
        """Stability: to_dict must NOT leak violations or schema_issues."""
        err = _pk_dup_error()
        d = err.to_dict()
        assert "violations" not in d
        assert "schema_issues" not in d
        assert "detail_rows" not in d  # old field removed
